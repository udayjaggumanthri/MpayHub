"""Parse BillAvenue MDM Excel sheets (blr_id / blr_name / category / coverage)."""

from __future__ import annotations

import re
from io import BytesIO
from typing import Any, BinaryIO

from openpyxl import load_workbook

BILLER_ID_RE = re.compile(r'^[A-Za-z0-9\-_]+$')
MAX_ROWS = 50_000
MAX_FILE_BYTES = 10 * 1024 * 1024

_ID_ALIASES = {'blr_id', 'biller_id', 'billerid', 'id'}
_NAME_ALIASES = {'blr_name', 'biller_name', 'billername', 'name'}
_CAT_ALIASES = {'blr_category_name', 'biller_category', 'billercategory', 'category', 'blr_category'}
_COV_ALIASES = {'blr_coverage', 'biller_coverage', 'coverage'}


def _norm_header(value: Any) -> str:
    return str(value or '').strip().lower().replace(' ', '_').replace('-', '_')


def _map_headers(headers: list[Any]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, raw in enumerate(headers):
        key = _norm_header(raw)
        if key in _ID_ALIASES and 'biller_id' not in mapping:
            mapping['biller_id'] = idx
        elif key in _NAME_ALIASES and 'biller_name' not in mapping:
            mapping['biller_name'] = idx
        elif key in _CAT_ALIASES and 'biller_category' not in mapping:
            mapping['biller_category'] = idx
        elif key in _COV_ALIASES and 'biller_coverage' not in mapping:
            mapping['biller_coverage'] = idx
    return mapping


def parse_mdm_excel(file_obj: BinaryIO | bytes, *, filename: str = '') -> list[dict[str, str]]:
    """
    Return deduped rows: biller_id, biller_name, biller_category, biller_coverage.

    Raises ValueError on invalid/empty sheets.
    """
    if isinstance(file_obj, (bytes, bytearray)):
        raw = bytes(file_obj)
        stream: BinaryIO = BytesIO(raw)
        size = len(raw)
    else:
        pos = file_obj.tell() if hasattr(file_obj, 'tell') else 0
        data = file_obj.read()
        size = len(data) if isinstance(data, (bytes, bytearray)) else 0
        if hasattr(file_obj, 'seek'):
            try:
                file_obj.seek(pos)
            except Exception:
                pass
        stream = BytesIO(data if isinstance(data, (bytes, bytearray)) else b'')

    if size > MAX_FILE_BYTES:
        raise ValueError(f'Excel file too large (max {MAX_FILE_BYTES // (1024 * 1024)}MB).')

    name = str(filename or '').lower()
    if name and not (name.endswith('.xlsx') or name.endswith('.xlsm') or name.endswith('.xls')):
        raise ValueError('Only Excel files (.xlsx) are supported.')

    try:
        wb = load_workbook(stream, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f'Unable to read Excel file: {exc}') from exc

    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration as exc:
            raise ValueError('Excel sheet is empty.') from exc

        colmap = _map_headers(list(header_row or []))
        if 'biller_id' not in colmap:
            raise ValueError('Missing biller ID column (expected blr_id or biller_id).')

        out: list[dict[str, str]] = []
        seen: set[str] = set()
        invalid = 0
        for row in rows_iter:
            if not row:
                continue
            if len(out) >= MAX_ROWS:
                raise ValueError(f'Excel has more than {MAX_ROWS} biller rows.')
            cells = list(row)
            bid_idx = colmap['biller_id']
            bid = str((cells[bid_idx] if bid_idx < len(cells) else '') or '').strip()
            if not bid:
                continue
            if not BILLER_ID_RE.match(bid):
                invalid += 1
                continue
            if bid in seen:
                continue
            seen.add(bid)

            def _cell(key: str) -> str:
                idx = colmap.get(key)
                if idx is None or idx >= len(cells):
                    return ''
                return str(cells[idx] or '').strip()

            out.append(
                {
                    'biller_id': bid,
                    'biller_name': _cell('biller_name'),
                    'biller_category': _cell('biller_category'),
                    'biller_coverage': _cell('biller_coverage'),
                }
            )
    finally:
        wb.close()

    if not out:
        raise ValueError('No valid biller IDs found in Excel.')
    if invalid and not out:
        raise ValueError(f'All {invalid} biller ID(s) had invalid format.')
    return out
